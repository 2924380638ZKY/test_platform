import json
import uuid
from datetime import datetime
import pandas as pd
import requests
from bson import ObjectId
from flask import Blueprint, request
from flask_restful import Api, Resource
from openpyxl.workbook import Workbook
from application import db
from application.account import login_authority
from application.account.login_authority import token_auth
from application.conf.configuration import platform_upload_url, platform_downLoad_url

collection = db["abilityCase"]
abilityCase = Blueprint("abilityCase", __name__, url_prefix="/api/v1/abilityCase")
api = Api(abilityCase)
log = db["log"]
model = db["model"]


# 添加功能用例
class Add(Resource):
    @token_auth
    def post(self):
        data = request.get_json()
        data["lastModifiedTime"] = str(datetime.now())
        # 实时获取当前操作token的用户信息
        data["lastModifiedBy"] = login_authority.user_data["data"]["account"]
        data_id = collection.insert_one(data).inserted_id
        # 日志打桩
        if data_id is not None:
            log.insert_one(
                {"optUserId": login_authority.user_data["data"]["userId"],
                 "optUserName": login_authority.user_data["data"]["username"],
                 "optAccount": login_authority.user_data["data"]["account"],
                 "optModule": "功能用例管理", "message": "添加功能用例", "optTime": str(datetime.now()),
                 "ip": request.remote_addr, "optSuccess": "True",
                 "roleId": login_authority.user_data["data"]["roleInfos"][0]["roleId"],
                 "roleName": login_authority.user_data["data"]["roleInfos"][0]["roleName"]})
            return dict(code=0, message="操作成功")
        log.insert_one(
            {"optUserId": login_authority.user_data["data"]["userId"],
             "optUserName": login_authority.user_data["data"]["username"],
             "optAccount": login_authority.user_data["data"]["account"],
             "optModule": "功能用例管理", "message": "添加功能用例", "optTime": str(datetime.now()),
             "ip": request.remote_addr, "optSuccess": "False",
             "roleId": login_authority.user_data["data"]["roleInfos"][0]["roleId"],
             "roleName": login_authority.user_data["data"]["roleInfos"][0]["roleName"]})
        return dict(code=1, messages="新增失败")


# 编辑功能用例
class Edit(Resource):
    @token_auth
    def post(self):
        data = request.get_json()
        result = collection.update_one({"_id": ObjectId(data["id"])},
                                       {"$set": {"caseName": data["caseName"], "abilityModelId": data["abilityModelId"],
                                                 "priority": data["priority"], "testData": data["testData"],
                                                 "expectResult": data["expectResult"],
                                                 "testingProcedure": data["testingProcedure"],
                                                 "preconditions": data["preconditions"],
                                                 "lastModifiedTime": str(datetime.now()), "lastModifiedBy": login_authority.user_data["data"]["account"]}})
        # 影响的数据条数
        count = result.modified_count
        if count > 0:
            log.insert_one(
                {"optUserId": login_authority.user_data["data"]["userId"],
                 "optUserName": login_authority.user_data["data"]["username"],
                 "optAccount": login_authority.user_data["data"]["account"],
                 "optModule": "功能用例管理", "message": "编辑功能用例", "optTime": str(datetime.now()),
                 "ip": request.remote_addr, "optSuccess": "True",
                 "roleId": login_authority.user_data["data"]["roleInfos"][0]["roleId"],
                 "roleName": login_authority.user_data["data"]["roleInfos"][0]["roleName"]})
            return dict(code=0, message="操作成功")
        log.insert_one(
            {"optUserId": login_authority.user_data["data"]["userId"],
             "optUserName": login_authority.user_data["data"]["username"],
             "optAccount": login_authority.user_data["data"]["account"],
             "optModule": "功能用例管理", "message": "编辑功能用例", "optTime": str(datetime.now()),
             "ip": request.remote_addr, "optSuccess": "False",
             "roleId": login_authority.user_data["data"]["roleInfos"][0]["roleId"],
             "roleName": login_authority.user_data["data"]["roleInfos"][0]["roleName"]})
        return dict(code=1, messages="编辑失败")


# 删除功能用例
class Delete(Resource):
    @token_auth
    def post(self):
        data = request.get_json()
        result = collection.delete_one({"_id": ObjectId(data["id"])})
        count = result.deleted_count
        if count > 0:
            log.insert_one(
                {"optUserId": login_authority.user_data["data"]["userId"],
                 "optUserName": login_authority.user_data["data"]["username"],
                 "optAccount": login_authority.user_data["data"]["account"],
                 "optModule": "功能用例管理", "message": "删除功能用例", "optTime": str(datetime.now()),
                 "ip": request.remote_addr, "optSuccess": "True",
                 "roleId": login_authority.user_data["data"]["roleInfos"][0]["roleId"],
                 "roleName": login_authority.user_data["data"]["roleInfos"][0]["roleName"]})
            return dict(code=0, message="操作成功")
        log.insert_one(
            {"optUserId": login_authority.user_data["data"]["userId"],
             "optUserName": login_authority.user_data["data"]["username"],
             "optAccount": login_authority.user_data["data"]["account"],
             "optModule": "功能用例管理", "message": "删除功能用例", "optTime": str(datetime.now()),
             "ip": request.remote_addr, "optSuccess": "False",
             "roleId": login_authority.user_data["data"]["roleInfos"][0]["roleId"],
             "roleName": login_authority.user_data["data"]["roleInfos"][0]["roleName"]})
        return dict(code=1, messages="删除失败")


# 获取功能用例列表
class Getlist(Resource):
    @token_auth
    def post(self):
        page = request.json.get("page")
        pageSize = request.json.get("pageSize")
        abilityModelId = request.json.get("abilityModelId")
        caseName = request.json.get("caseName")
        priority = request.json.get("priority")
        query = {}
        if abilityModelId:
            if len(abilityModelId) == 1:
                query["abilityModelId"] = abilityModelId[0]
            elif len(abilityModelId) == 2:
                query["abilityModelId"] = {"$all": abilityModelId[:2]}
            else:
                query["abilityModelId"] = abilityModelId
        if caseName:
            # 模糊查询，匹配区分大小写
            query["caseName"] = {"$regex": caseName, "$options": "i"}
        # 在python里对0判断是为false的，所以加上一个or，P0存在库里是表示为0的
        if priority or priority == 0:
            query["priority"] = priority
        # 跳过前N条记录，限制只展示pagesize条记录
        results = collection.find(query).skip((page - 1) * pageSize).limit(pageSize)
        data = [{"id": str(result["_id"]), "abilityModelId": result["abilityModelId"], "caseName": result["caseName"],
                 "priority": result["priority"], "preconditions": result["preconditions"],
                 "testingProcedure": result["testingProcedure"],
                 "testData": result["testData"], "expectResult": result["expectResult"],
                 "lastModifiedTime": result["lastModifiedTime"], "lastModifiedBy": result["lastModifiedBy"]
                 } for result in results]
        total_count = collection.count_documents(query)
        total_page = (total_count + pageSize - 1) // pageSize
        response_data = {
            "currentPage": page,
            "currentPageSize": pageSize,
            "totalPage": total_page,
            "totalCount": total_count,
            "dataList": data
        }
        return dict(code=0, message="操作成功", data=response_data)


# 导入功能用例
class Import(Resource):
    @token_auth
    def post(self):
        file = request.files["file"]
        # 获取用例数据和用例表名
        excel_data = pd.read_excel(file, sheet_name=None, engine='openpyxl')
        excel_name = file.filename
        excel_name = excel_name.strip(".xlsx")
        # 根据我们的定义，定义excel表的表名为模块的最外层级，模块的名称；excel表的工作表名为维护里的第二层级；里面数据的功能模块为第三层级
        # 如果这个数据库里没有为这个名称的模块，先生成
        if not model.find_one({"modelName": excel_name}):
            model.insert_one({"modelName": excel_name, "desc": "", "modelList": []})

        # 将excel数据存储到数据库中
        for sheet_name, sheet_data in excel_data.items():
            # 判断第二层级是否存在
            second_label_exists = False
            second_sheet = model.find_one({"modelName": excel_name})
            for item1 in second_sheet["modelList"]:
                if item1.get("label") == sheet_name:
                    second_label_exists = True
                    break
            # 如果不存在，生成第二层级的结构体，并append到对应的模块数据里
            if not second_label_exists:
                second_data = {
                    "id": str(uuid.uuid4()),
                    "label": sheet_name,
                    "parentId": "",
                    "level": 1,
                    "children": []
                }
                second_sheet["modelList"].append(second_data)
                model.update_one({"modelName": excel_name},
                                 {"$set": {"modelList": second_sheet["modelList"]}})
            # 将dataframe转换为字典列表
            data_list = sheet_data.to_dict("records")
            # 将数据插入到mongodb数据库中
            for single_data in data_list:
                third_label_exists = False
                # 我们原本的excel表里时有这几个字段，但是存到库和测试平台的时候不需要这些字段，直接pop去掉
                if "用例编号" in single_data:
                    single_data.pop("用例编号")
                if "子功能模块" in single_data:
                    single_data.pop("子功能模块")
                if "实际结果" in single_data:
                    single_data.pop("实际结果")
                if "备注" in single_data:
                    single_data.pop("备注")
                # 修改字段名称，修改优先级为int类型
                single_data["abilityModelId"] = single_data.pop("功能模块")
                single_data["priority"] = single_data.pop("优先级")
                if single_data["priority"] == "P0":
                    single_data["priority"] = 0
                elif single_data["priority"] == "P1":
                    single_data["priority"] = 1
                elif single_data["priority"] == "P2":
                    single_data["priority"] = 2
                else:
                    single_data["priority"] = 3
                single_data["caseName"] = single_data.pop("用例名称")
                single_data["preconditions"] = single_data.pop("前置条件")
                single_data["testingProcedure"] = single_data.pop("测试步骤")
                single_data["testData"] = single_data.pop("测试数据")
                single_data["expectResult"] = single_data.pop("预期结果")
                # 判断第三层级是否存在
                for item in second_sheet["modelList"]:
                    if item.get("label") == sheet_name:
                        for item1 in item["children"]:
                            if item1.get("label") == single_data["abilityModelId"]:
                                third_label_exists = True
                                break
                        # 不存在的话，生成第三层级的结构体，并append到对应的模块数据里
                        if not third_label_exists:
                            third_data = {
                                "id": str(uuid.uuid4()),
                                "label": single_data["abilityModelId"],
                                "parentId": "",
                                "level": 2,
                                "children": []
                            }
                            item["children"].append(third_data)
                            model.update_one({"modelName": excel_name},
                                             {"$set": {"modelList": second_sheet["modelList"]}})
                # 生成功能模块的id数组并赋值
                x = model.find_one({"modelName": excel_name})
                number_id = []
                number1 = str(x["_id"])
                number_id.append(number1)
                for item in x["modelList"]:
                    if item.get("label") == sheet_name:
                        number2 = item["id"]
                        number_id.append(number2)
                        for item1 in item["children"]:
                            if item1.get("label") == single_data["abilityModelId"]:
                                number3 = item1["id"]
                                number_id.append(number3)
                single_data["abilityModelId"] = number_id
                single_data["lastModifiedTime"] = str(datetime.now())
                single_data["lastModifiedBy"] = login_authority.user_data["data"]["account"]
            # 插入数据到库里
            collection.insert_many(data_list)
        return dict(code=0, message="操作成功")


# 导出用例
class Export(Resource):
    @token_auth
    def post(self):
        abilityModelId = request.json.get("abilityModelId")
        caseName = request.json.get("caseName")
        priority = request.json.get("priority")
        query = {}
        if abilityModelId:
            if len(abilityModelId) == 1:
                query["abilityModelId"] = abilityModelId[0]
            elif len(abilityModelId) == 2:
                query["abilityModelId"] = {"$all": abilityModelId[:2]}
            else:
                query["abilityModelId"] = abilityModelId
        if caseName:
            query["caseName"] = {"$regex": caseName, "$options": "i"}
        if priority or priority == 0:
            query["priority"] = priority
        results = collection.find(query)
        # 获取用例的id列表
        data = []
        for result in results:
            data.append(str(result["_id"]))
        second_label_name = ''
        third_label_name = ''
        wb = Workbook()
        ws_dict = {}
        ws_number = {}
        # 获取model数据，获取的数据为该id下的所有信息，因为都是同一个大模块下的数据，所以随便找一个数据获取"abilityModelId"的第一个id，都是一样的。
        file_name_number = collection.find_one({"_id": ObjectId(data[0])})
        model_data = model.find_one({"_id": ObjectId(file_name_number["abilityModelId"][0])})
        # 功能模块的第二层级；没被使用，就生成工作表。并且在ws_dict里生成一个键值对，标明已经生成过了
        for item in data:
            case = collection.find_one({"_id": ObjectId(item)})
            ability_model_id = case['abilityModelId'][1]
            if ability_model_id not in ws_dict:
                ws_dict[ability_model_id] = wb.create_sheet(title=f'Sheet{ability_model_id}')
                ws_number[ability_model_id] = 1
            ws = ws_dict[ability_model_id]
            # 在当前工作表里生成首行
            ws.cell(row=1, column=1, value='用例编号')
            ws.cell(row=1, column=2, value='功能模块')
            ws.cell(row=1, column=3, value='优先级')
            ws.cell(row=1, column=4, value='用例名称')
            ws.cell(row=1, column=5, value='前置条件')
            ws.cell(row=1, column=6, value='测试步骤')
            ws.cell(row=1, column=7, value='测试数据')
            ws.cell(row=1, column=8, value='预期结果')
            # 获取第二层级的名称
            for item1 in model_data["modelList"]:
                if item1.get("id") == (case['abilityModelId'][1]):
                    second_label_name = item1.get("label")
                    break
            # 修改工作表的名称，修改序号和优先级
            ws.title = second_label_name
            case["_id"] = ws_number[ability_model_id]
            if case["priority"] == 0:
                case["priority"] = "P0"
            elif case["priority"] == 1:
                case["priority"] = "P1"
            elif case["priority"] == 2:
                case["priority"] = "P2"
            else:
                case["priority"] = "P3"
            # 生成第三层级的名称
            for item1 in model_data["modelList"]:
                if item1.get("id") == (case['abilityModelId'][1]):
                    for item2 in item1["children"]:
                        if item2.get("id") == (case['abilityModelId'][2]):
                            third_label_name = item2.get("label")
                            break
            case["abilityModelId"] = third_label_name
            # 去掉操作事件和操作人
            case.pop("lastModifiedTime")
            case.pop("lastModifiedBy")
            # 写入数据，将序号加1
            row = [str(value) for value in case.values()]
            ws.append(row)
            ws_number[ability_model_id] = ws_number[ability_model_id] + 1

        # 修改文件名
        ws = wb['Sheet']
        wb.remove(ws)
        file_name = (model_data["modelName"] + ".xlsx")
        wb.save(r"application/file_storage/" + file_name)

        # 上传到文件服务器，返回下载地址
        file = {'file': open(r"application/file_storage/" + file_name, 'rb')}
        upload_file = requests.request(method="post", url=platform_upload_url, files=file)
        info = json.loads(upload_file.text)
        downLoad_data = {"objectId": info["data"]["id"]}
        download_address = requests.request(method="post",
                                            url=platform_downLoad_url,
                                            json=downLoad_data)

        return dict(code=0, data=json.loads(download_address.text)["data"])


# 模板下载
class downLoad(Resource):
    @token_auth
    def post(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='用例编号')
        ws.cell(row=1, column=2, value='功能模块')
        ws.cell(row=1, column=3, value='优先级')
        ws.cell(row=1, column=4, value='用例名称')
        ws.cell(row=1, column=5, value='前置条件')
        ws.cell(row=1, column=6, value='测试步骤')
        ws.cell(row=1, column=7, value='测试数据')
        ws.cell(row=1, column=8, value='预期结果')
        file_name = "导入模板.xlsx"
        wb.save(r"application/file_storage/" + file_name)
        file = {'file': open(r"application/file_storage/" + file_name, 'rb')}
        upload_file = requests.request(method="post", url=platform_upload_url, files=file)
        info = json.loads(upload_file.text)
        downLoad_data = {"objectId": info["data"]["id"]}
        download_address = requests.request(method="post",
                                            url=platform_downLoad_url,
                                            json=downLoad_data)
        return dict(code=0, data=json.loads(download_address.text)["data"])


api.add_resource(Add, "/addAbilityCase", endpoint="addAbilityCase")
api.add_resource(Edit, "/editAbilityCase", endpoint="editAbilityCase")
api.add_resource(Delete, "/deleteAbilityCase", endpoint="deleteAbilityCase")
api.add_resource(Getlist, "/getAbilityCaseList", endpoint="getAbilityCaseList")
api.add_resource(Import, "/importAbilityCaseList", endpoint="importAbilityCaseList")
api.add_resource(Export, "/exportAbilityCaseList", endpoint="exportAbilityCaseList")
api.add_resource(downLoad, "/downLoadTemplate", endpoint="downLoadTemplate")
