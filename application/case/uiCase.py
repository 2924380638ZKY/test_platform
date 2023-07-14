import json
from datetime import datetime

import requests
from bson import ObjectId
from flask import Blueprint, request
from flask_restful import Api, Resource
from openpyxl.workbook import Workbook

from application import db
from application.account import login_authority
from application.account.login_authority import token_auth
from application.conf.configuration import platform_upload_url, platform_downLoad_url

collection = db["uiCase"]
uiCase = Blueprint("uiCase", __name__, url_prefix="/api/v1/uiCase")
api = Api(uiCase)
log = db["log"]
kit = db['kit']
event = db['event']
model = db["model"]


# 新增UI用例
class Add(Resource):
    @token_auth
    def post(self):
        data = request.get_json()
        data["eventName"] = []
        for eventid in data["eventStep"]:
            result = event.find_one({"_id": ObjectId(eventid["eventId"])})
            data["eventName"].append(result["eventName"])
        data["lastModifiedTime"] = str(datetime.now())
        data["lastModifiedBy"] = login_authority.user_data["data"]["account"]
        data_id = collection.insert_one(data).inserted_id
        if data_id is not None:
            log.insert_one(
                {"optUserId": login_authority.user_data["data"]["userId"],
                 "optUserName": login_authority.user_data["data"]["username"],
                 "optAccount": login_authority.user_data["data"]["account"],
                 "optModule": "UI用例管理", "message": "添加UI用例", "optTime": str(datetime.now()),
                 "ip": request.remote_addr, "optSuccess": "True",
                 "roleId": login_authority.user_data["data"]["roleInfos"][0]["roleId"],
                 "roleName": login_authority.user_data["data"]["roleInfos"][0]["roleName"]})
            return dict(code=0, message="操作成功")
        log.insert_one(
            {"optUserId": login_authority.user_data["data"]["userId"],
             "optUserName": login_authority.user_data["data"]["username"],
             "optAccount": login_authority.user_data["data"]["account"],
             "optModule": "UI用例管理", "message": "添加UI用例", "optTime": str(datetime.now()),
             "ip": request.remote_addr, "optSuccess": "False",
             "roleId": login_authority.user_data["data"]["roleInfos"][0]["roleId"],
             "roleName": login_authority.user_data["data"]["roleInfos"][0]["roleName"]})
        return dict(code=1, messages="新增失败")


# 编辑UI用例
class Edit(Resource):
    @token_auth
    def post(self):
        data = request.get_json()
        data["eventName"] = []
        for eventid in data["eventStep"]:
            result = event.find_one({"_id": ObjectId(eventid["eventId"])})
            data["eventName"].append(result["eventName"])
        result1 = collection.update_one({"_id": ObjectId(data["id"])},
                                        {"$set": {"abilityModelId": data["abilityModelId"], "login": data["login"],
                                                  "priority": data["priority"],
                                                  "caseName": data["caseName"], "eventStep": data["eventStep"],
                                                  "eventName": data["eventName"],
                                                  "testData": data["testData"], "expectResult": data["expectResult"],
                                                  "lastModifiedTime": str(datetime.now()),
                                                  "lastModifiedBy": login_authority.user_data["data"]["account"]}})
        # 同步套件中的信息
        ui_kits = kit.find({"kitType": 2})
        for ui_kit in ui_kits:
            for ui_case in ui_kit['caseList']:
                if ui_case['id'] == data['id']:
                    kit.update_one({'caseList.id': data['id']},
                                   {'$set': {'caseList.$.login': data["login"],
                                             'caseList.$.abilityModelId': data["abilityModelId"],
                                             'caseList.$.caseName': data["caseName"],
                                             'caseList.$.eventStep': data["eventStep"],
                                             'caseList.$.eventName': data["eventName"],
                                             'caseList.$.priority': data["priority"],
                                             'caseList.$.expectResult': data["expectResult"],
                                             'caseList.$.testData': data["testData"]}})
        # 影响的数据条数
        count = result1.modified_count
        if count > 0:
            log.insert_one(
                {"optUserId": login_authority.user_data["data"]["userId"],
                 "optUserName": login_authority.user_data["data"]["username"],
                 "optAccount": login_authority.user_data["data"]["account"],
                 "optModule": "UI用例管理", "message": "编辑UI用例", "optTime": str(datetime.now()),
                 "ip": request.remote_addr, "optSuccess": "True",
                 "roleId": login_authority.user_data["data"]["roleInfos"][0]["roleId"],
                 "roleName": login_authority.user_data["data"]["roleInfos"][0]["roleName"]})
            return dict(code=0, message="操作成功")
        log.insert_one(
            {"optUserId": login_authority.user_data["data"]["userId"],
             "optUserName": login_authority.user_data["data"]["username"],
             "optAccount": login_authority.user_data["data"]["account"],
             "optModule": "UI用例管理", "message": "编辑UI用例", "optTime": str(datetime.now()),
             "ip": request.remote_addr, "optSuccess": "False",
             "roleId": login_authority.user_data["data"]["roleInfos"][0]["roleId"],
             "roleName": login_authority.user_data["data"]["roleInfos"][0]["roleName"]})
        return dict(code=1, messages="编辑失败")


# 删除UI用例，如果用例被套件使用了，删除失败
class Delete(Resource):
    @token_auth
    def post(self):
        data = request.get_json()
        used_in_kit = kit.find_one({"caseList.id": data["id"]})
        if used_in_kit:
            log.insert_one(
                {"optUserId": login_authority.user_data["data"]["userId"],
                 "optUserName": login_authority.user_data["data"]["username"],
                 "optAccount": login_authority.user_data["data"]["account"],
                 "optModule": "UI用例管理", "message": "删除UI用例", "optTime": str(datetime.now()),
                 "ip": request.remote_addr, "optSuccess": "False",
                 "roleId": login_authority.user_data["data"]["roleInfos"][0]["roleId"],
                 "roleName": login_authority.user_data["data"]["roleInfos"][0]["roleName"]})
            return dict(code=10, message="该UI用例正在被套件使用，删除失败")

        result = collection.delete_one({"_id": ObjectId(data["id"])})
        count = result.deleted_count
        if count > 0:
            log.insert_one(
                {"optUserId": login_authority.user_data["data"]["userId"],
                 "optUserName": login_authority.user_data["data"]["username"],
                 "optAccount": login_authority.user_data["data"]["account"],
                 "optModule": "UI用例管理", "message": "删除UI用例", "optTime": str(datetime.now()),
                 "ip": request.remote_addr, "optSuccess": "True",
                 "roleId": login_authority.user_data["data"]["roleInfos"][0]["roleId"],
                 "roleName": login_authority.user_data["data"]["roleInfos"][0]["roleName"]})
            return dict(code=0, message="操作成功")
        log.insert_one(
            {"optUserId": login_authority.user_data["data"]["userId"],
             "optUserName": login_authority.user_data["data"]["username"],
             "optAccount": login_authority.user_data["data"]["account"],
             "optModule": "UI用例管理", "message": "删除UI用例", "optTime": str(datetime.now()),
             "ip": request.remote_addr, "optSuccess": "False",
             "roleId": login_authority.user_data["data"]["roleInfos"][0]["roleId"],
             "roleName": login_authority.user_data["data"]["roleInfos"][0]["roleName"]})
        return dict(code=1, messages="删除失败")


# 获取某个UI用例的信息
class Get(Resource):
    @token_auth
    def post(self):
        data = request.get_json()
        result = collection.find_one({"_id": ObjectId(data["id"])})
        if result is not None:
            result1 = {"id": str(result["_id"]), "abilityModelId": result["abilityModelId"],
                       "priority": result["priority"], "login": data["login"],
                       "caseName": result["caseName"], "eventStep": result["eventStep"],
                       "testData": result["testData"], "expectResult": result["expectResult"]}
            return dict(code=0, message="操作成功", data=result1)
        return dict(code=1, messages="获取信息失败")


# 获取UI用例列表
class Getlist(Resource):
    @token_auth
    def post(self):
        page = request.json.get("page")
        pageSize = request.json.get("pageSize")
        abilityModelId = request.json.get("abilityModelId")
        caseName = request.json.get("caseName")
        priority = request.json.get("priority")
        query = {}
        if abilityModelId:
            if len(abilityModelId) == 1:
                query["abilityModelId"] = abilityModelId[0]
            elif len(abilityModelId) == 2:
                query["abilityModelId"] = {"$all": abilityModelId[:2]}
            elif len(abilityModelId) == 3:
                query["abilityModelId"] = abilityModelId
        if caseName:
            query["caseName"] = {"$regex": caseName, "$options": "i"}  # 模糊查询，匹配区分大小写
        if priority or priority == 0:
            query["priority"] = priority
        results = collection.find(query).skip((page - 1) * pageSize).limit(pageSize)  # 跳过前N条记录，限制只展示pagesize条记录
        data = [{"id": str(result["_id"]), "abilityModelId": result["abilityModelId"], "login": result["login"],
                 "caseName": result["caseName"],
                 "priority": result["priority"], "eventStep": result["eventStep"], "eventName": result["eventName"],
                 "testData": result["testData"],
                 "expectResult": result["expectResult"],
                 "lastModifiedTime": result["lastModifiedTime"], "lastModifiedBy": result["lastModifiedBy"]
                 } for result in results]
        total_count = collection.count_documents(query)
        total_page = (total_count + pageSize - 1) // pageSize
        response_data = {
            "currentPage": page,
            "currentPageSize": pageSize,
            "totalPage": total_page,
            "totalCount": total_count,
            "dataList": data
        }
        return dict(code=0, message="操作成功", data=response_data)


# 导出UI用例
class Export(Resource):
    @token_auth
    def post(self):
        abilityModelId = request.json.get("abilityModelId")
        caseName = request.json.get("caseName")
        priority = request.json.get("priority")
        query = {}
        if abilityModelId:
            if len(abilityModelId) == 1:
                query["abilityModelId"] = abilityModelId[0]
            elif len(abilityModelId) == 2:
                query["abilityModelId"] = {"$all": abilityModelId[:2]}
            else:
                query["abilityModelId"] = abilityModelId
        if caseName:
            query["caseName"] = {"$regex": caseName, "$options": "i"}  # 模糊查询，匹配区分大小写
        if priority or priority == 0:
            query["priority"] = priority
        results = collection.find(query)
        data = []
        for result in results:
            data.append(str(result["_id"]))
        second_label_name = ''
        third_label_name = ''
        wb = Workbook()
        ws_dict = {}
        ws_number = {}
        # 获取model数据，获取的数据为该id下的所有信息
        file_name_number = collection.find_one({"_id": ObjectId(data[0])})
        model_data = model.find_one({"_id": ObjectId(file_name_number["abilityModelId"][0])})

        for item in data:

            case = collection.find_one({"_id": ObjectId(item)})
            ability_model_id = case['abilityModelId'][1]
            if ability_model_id not in ws_dict:
                ws_dict[ability_model_id] = wb.create_sheet(title=f'Sheet{ability_model_id}')
                ws_number[ability_model_id] = 1
            ws = ws_dict[ability_model_id]

            ws.cell(row=1, column=1, value='用例编号')
            ws.cell(row=1, column=2, value='功能模块')
            ws.cell(row=1, column=3, value='优先级')
            ws.cell(row=1, column=4, value='用例名称')
            ws.cell(row=1, column=5, value='是否需要登录')
            ws.cell(row=1, column=6, value='事件步骤')
            ws.cell(row=1, column=7, value='测试数据')
            ws.cell(row=1, column=8, value='预期结果')

            for item1 in model_data["modelList"]:
                if item1.get("id") == (case['abilityModelId'][1]):
                    second_label_name = item1.get("label")
                    break
            # 写入数据
            ws.title = second_label_name
            case["_id"] = ws_number[ability_model_id]
            if case["priority"] == 0:
                case["priority"] = "P0"
            elif case["priority"] == 1:
                case["priority"] = "P1"
            elif case["priority"] == 2:
                case["priority"] = "P2"
            else:
                case["priority"] = "P3"

            case["eventStep"] = case["eventName"]

            for item1 in model_data["modelList"]:
                if item1.get("id") == (case['abilityModelId'][1]):
                    for item2 in item1["children"]:
                        if item2.get("id") == (case['abilityModelId'][2]):
                            third_label_name = item2.get("label")
                            break

            case["abilityModelId"] = third_label_name
            case.pop("lastModifiedTime")
            case.pop("lastModifiedBy")
            if 'eventName' in case:
                case.pop("eventName")
            row = [str(value) for value in case.values()]
            ws.append(row)
            ws_number[ability_model_id] = ws_number[ability_model_id] + 1

        # 修改文件名
        ws = wb['Sheet']
        wb.remove(ws)
        file_name = (model_data["modelName"] + "UI" + ".xlsx")
        wb.save(r"application/file_storage/" + file_name)

        # 上传到文件服务器，返回下载地址
        file = {'file': open(r"application/file_storage/" + file_name, 'rb')}
        upload_file = requests.request(method="post", url=platform_upload_url, files=file)
        info = json.loads(upload_file.text)
        downLoad_data = {"objectId": info["data"]["id"]}
        download_address = requests.request(method="post",
                                            url=platform_downLoad_url,
                                            json=downLoad_data)

        return dict(code=0, data=json.loads(download_address.text)["data"])


api.add_resource(Add, "/addUiCase", endpoint="addUiCase")
api.add_resource(Edit, "/editUiCase", endpoint="editUiCase")
api.add_resource(Delete, "/deleteUiCase", endpoint="deleteUiCase")
api.add_resource(Get, "/getUiCaseInfo", endpoint="getUiCaseInfo")
api.add_resource(Getlist, "/getCaseList", endpoint="getCaseList")
api.add_resource(Export, "/exportUiCaseList", endpoint="exportUiCaseList")
